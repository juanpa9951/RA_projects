####  This is New camaras V0 for MULTIPLE STACKS at the time

import cv2
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import os
import csv
from datetime import datetime
import threading
import pandas as pd
from openpyxl import load_workbook

# Global variables for camera
capture_flag = True
camera_feed = None

# Directory for saving screenshots and data
# screenshot_dir = r"C:\Users\Terminal6\Rewair A S\Terminal1 - RAES013 - ASM003"
# data_dir = r"C:\Users\Terminal6\OneDrive - Rewair A S\Laser - ASM003\00_NO TOCAR"
screenshot_dir = "screenshots"
data_dir = "data"


# Ensure directories exist
os.makedirs(screenshot_dir, exist_ok=True)
os.makedirs(data_dir, exist_ok=True)

def load_data(file_path, sheet_name="Sheet1"):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header
        data.append({
            "Project": row[0],
            "Side": row[1],
            "Packing": row[2],
            "Name": row[3]
        })
    return data

# Update "Side" and "Packing" comboboxes based on "Project"
def update_side_and_packing():
    project_filter = list2.get()

    # Filter unique sides and packings based on project
    filtered_sides = sorted(set(
        row["Side"] for row in data
        if project_filter == "" or row["Project"] == project_filter
    ))
    filtered_packings = sorted(set(
        row["Packing"] for row in data
        if project_filter == "" or row["Project"] == project_filter
    ))

    # Update the "Side" and "Packing" comboboxes
    list5["values"] = filtered_sides
    list3["values"] = filtered_packings
    list5.set("")  # Reset the current selection
    list3.set("")  # Reset the current selection

    # Clear and reset "Name" options
    update_name_options()

# Update "Name" combobox based on filters
def update_name_options():
    project_filter = list2.get()
    side_filter = list5.get()
    packing_filter = list3.get()

    # Filter the data
    filtered_names = [
        row["Name"] for row in data
        if (project_filter == "" or row["Project"] == project_filter)
           and (side_filter == "" or row["Side"] == side_filter)
           and (packing_filter == "" or row["Packing"] == packing_filter)
    ]

    # Update "Name" combobox
    list6["values"] = filtered_names
    list6.set("")  # Reset the combobox value
    list7["values"] = filtered_names
    list7.set("")  # Reset the combobox value
    list8["values"] = filtered_names
    list8.set("")  # Reset the combobox value
    list9["values"] = filtered_names
    list9.set("")  # Reset the combobox value
    list10["values"] = filtered_names
    list10.set("")  # Reset the combobox value
    list11["values"] = filtered_names
    list11.set("")  # Reset the combobox value
    list12["values"] = filtered_names
    list12.set("")  # Reset the combobox value
    list13["values"] = filtered_names
    list13.set("")  # Reset the combobox value


# Load the Excel file data
# file_path = r"C:\Users\Terminal6\PycharmProjects\PythonProject\nombres_proyectos.xlsx"  # Replace with your Excel file path
file_path = "nombres_proyectos.xlsx"  # Replace with your Excel file path
data = load_data(file_path)


# Function to save screenshot and data
def save_screenshot_and_data():
    global camera_feed

    if camera_feed is None:
        print("Camera not initialized!")
        return

    # Collect user data from the entries
    user_data = {
        "ODF": entry1.get(),
        "Pala": list1.get(),
        "Proyect": list2.get(),
        "Packaging": list3.get(),
        "Side": list5.get(),
        "Stack1": list6.get(),
        "Stack2": list7.get(),
        "Stack3": list8.get(),
        "Stack4": list9.get(),
        "Stack5": list10.get(),
        "Stack6": list11.get(),
        "Stack7": list12.get(),
        "Stack8": list13.get(),
        "Maquina": "asm003",
        "FinCosido": "NaN"
    }

    # Sanitize inputs for filenames (remove spaces and special characters)
    sanitized_data = {key: "".join(e for e in value if e.isalnum() or e in "_-") for key, value in user_data.items()}

    # Create a unique filename using the entries
    filename_base = f"{sanitized_data['Stack1']}_{sanitized_data['Stack2']}_{sanitized_data['Stack3']}_{sanitized_data['Stack4']}_{sanitized_data['Stack5']}_{sanitized_data['Stack6']}_{sanitized_data['Stack7']}_{sanitized_data['Stack8']}"
    filename_base = filename_base[:100]  # Limit filename length for safety
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screenshot_filename = f"{filename_base}_{timestamp}.png"
    screenshot_path = os.path.join(screenshot_dir, screenshot_filename)
    print(f"Screenshot saved to: {screenshot_path}")

    # Prepare the image for saving with overlay
    image_with_text = camera_feed.copy()
    y_offset = 30  # Initial y-coordinate for text overlay
    for key, value in user_data.items():
        text = f"{key}: {value}"
        cv2.putText(image_with_text,text,(10, y_offset), cv2.FONT_HERSHEY_SIMPLEX,0.6,(0, 255, 0),2,cv2.LINE_AA,)
        y_offset += 30  # Move down for the next line of text

    # Save the image with overlay
    cv2.imwrite(screenshot_path, image_with_text)


    # check stack names
    stack_names=[]
    stack_names.append(sanitized_data['Stack1'])
    stack_names.append(sanitized_data['Stack2'])
    stack_names.append(sanitized_data['Stack3'])
    stack_names.append(sanitized_data['Stack4'])
    stack_names.append(sanitized_data['Stack5'])
    stack_names.append(sanitized_data['Stack6'])
    stack_names.append(sanitized_data['Stack7'])
    stack_names.append(sanitized_data['Stack8'])

    for name in stack_names:
        if name!="":
            user_data = {
                "ODF": entry1.get(),
                "Pala": list1.get(),
                "Proyect": list2.get(),
                "Packaging": list3.get(),
                "Side": list5.get(),
                "Stack": name,
                "Maquina": "asm003",
                "FinCosido": "NaN"
            }

            # Save data to a CSV file
            csv_file_path = os.path.join(data_dir, "user_data.csv")
            file_exists = os.path.isfile(csv_file_path)
            with open(csv_file_path, mode="a", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=list(user_data.keys()) + ["timestamp"])
                if not file_exists:
                    writer.writeheader()
                writer.writerow({**user_data, "timestamp": timestamp})

            # Update the status label with the filename
            status_label.config(text=f"Ultima foto guardada: {screenshot_filename}")
            cosido_label.config(text="")
            print("Data saved to CSV.")


# Function to smark fin cosido
def save_stiching():
    global camera_feed

    if camera_feed is None:
        print("Camera not initialized!")
        return

    # Collect user data from the entries
    user_data = {
        "ODF": entry1.get(),
        "Pala": list1.get(),
        "Proyect": list2.get(),
        "Packaging": list3.get(),
        "Side": list5.get(),
        "Stack1": list6.get(),
        "Stack2": list7.get(),
        "Stack3": list8.get(),
        "Stack4": list9.get(),
        "Stack5": list10.get(),
        "Stack6": list11.get(),
        "Stack7": list12.get(),
        "Stack8": list13.get(),
        "Maquina": "asm003",
        "FinCosido": "NaN"
    }

    # Sanitize inputs for filenames (remove spaces and special characters)
    sanitized_data = {key: "".join(e for e in value if e.isalnum() or e in "_-") for key, value in user_data.items()}
    #
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # check stack names
    stack_names=[]
    stack_names.append(sanitized_data['Stack1'])
    stack_names.append(sanitized_data['Stack2'])
    stack_names.append(sanitized_data['Stack3'])
    stack_names.append(sanitized_data['Stack4'])
    stack_names.append(sanitized_data['Stack5'])
    stack_names.append(sanitized_data['Stack6'])
    stack_names.append(sanitized_data['Stack7'])
    stack_names.append(sanitized_data['Stack8'])

    for name in stack_names:
        if name != "":
            user_data = {
                "ODF": entry1.get(),
                "Pala": list1.get(),
                "Proyect": list2.get(),
                "Packaging": list3.get(),
                "Side": list5.get(),
                "Stack": name,
                "Maquina": "asm003",
                "FinCosido": "FINALIZADO"
            }

            # Save data to a CSV file
            csv_file_path = os.path.join(data_dir, "user_data.csv")
            file_exists = os.path.isfile(csv_file_path)
            with open(csv_file_path, mode="a", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=list(user_data.keys()) + ["timestamp"])
                if not file_exists:
                    writer.writeheader()
                writer.writerow({**user_data, "timestamp": timestamp})

            # Update the status label with the filename
            stack_name = f"{sanitized_data['Stack1']}_{sanitized_data['Stack2']}_{sanitized_data['Stack3']}_{sanitized_data['Stack4']}_{sanitized_data['Stack5']}_{sanitized_data['Stack6']}_{sanitized_data['Stack7']}_{sanitized_data['Stack8']}"
            cosido_label.config(text=f"Registrado fin cosido STACK: {stack_name}")

            print("Data saved to CSV.")


# Function to run the live camera stream
def run_camera():
    global capture_flag, camera_feed

    # cap = cv2.VideoCapture("rtsp://LP003:LP003ASM@192.168.2.76:554/stream1")  ### asm001
    # cap = cv2.VideoCapture("rtsp://LP002:LP002ASM@192.168.2.72:554/stream1")  ### asm005
    # cap = cv2.VideoCapture("rtsp://MSM005:LP005ASM@172.16.58.15:554/stream1")  ### ud tapes
    cap = cv2.VideoCapture("rtsp://LP006:LP006ASM@192.168.2.79:554/stream1")  ### asm003 der
    # cap = cv2.VideoCapture("rtsp://LP005:LP005ASM@192.168.2.75:554/stream1")  ### asm003 izq
    # cap = cv2.VideoCapture("rtsp://LP001:LP001ASM@192.168.2.71:554/stream1")  ### asm002 izq
    # cap = cv2.VideoCapture("rtsp://LP004:LP004ASM@192.168.2.78:554/stream1")  ### asm002 der
    # cap = cv2.VideoCapture("rtsp://LP008:LP008ASM@192.168.2.82:554/stream1") ### asm004 der
    # cap = cv2.VideoCapture("rtsp://LP009:LP009ASM@192.168.2.84:554/stream1") ### asm004 izq
    # cap = cv2.VideoCapture("rtsp://RA-camara3:RewAir2023@172.16.58.16:554/stream1")  ## tagging 1
    # cap = cv2.VideoCapture("rtsp://RA-camaras:RewAir2023@172.16.58.142:554/stream1")  ## tagging 1 auxiliar
    # cap = cv2.VideoCapture("rtsp://RA-camara4:RewAir2023@172.16.58.17:554/stream1") ## tagging 2
    # cap = cv2.VideoCapture("rtsp://RA-camara2:RewAir2023@172.16.58.180:554/stream1")  ## tagging 2 auxiliar

    while capture_flag:
        ret, frame = cap.read()
        if ret:
            camera_feed = frame.copy()
            cv2.namedWindow("Live Camera Feed", cv2.WINDOW_NORMAL)
            cv2.imshow("Live Camera Feed", frame)
              #### CAREFUL HERE, THE 2ND ARGUMENT IS FOR ADAPTING TO ANY SCREEN SIZE
        if cv2.waitKey(1) & 0xFF == ord("q"):  # Press 'q' to exit
            capture_flag = False
            break

    cap.release()
    cv2.destroyAllWindows()

# Function to stop the program gracefully
def on_close():
    global capture_flag
    capture_flag = False
    root.destroy()




# GUI setup
root = tk.Tk()
root.title("Camera & Data Logger")
root.protocol("WM_DELETE_WINDOW", on_close)

# Extract unique values for the "Project" column
projects = sorted(set(row["Project"] for row in data))

# Add GUI elements

tk.Label(root, text="ODF:").grid(row=0, column=0, padx=5, pady=5)
entry1 = tk.Entry(root)
entry1.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Pala S:").grid(row=0, column=2, padx=5, pady=5)
list1 = ttk.Combobox(root, values=["1", "2", "3","4", "5", "6","7", "8", "9","10"])
list1.grid(row=0, column=3, padx=5, pady=5)

tk.Label(root, text="Proyect:").grid(row=1, column=0, padx=5, pady=5)
list2 = ttk.Combobox(root, values=projects, state="readonly")
list2.grid(row=1, column=1, padx=5, pady=5)
list2.bind("<<ComboboxSelected>>", lambda e: update_side_and_packing())

tk.Label(root, text="Side:").grid(row=1, column=2, padx=5, pady=5)
list5 = ttk.Combobox(root, state="readonly")
list5.grid(row=1, column=3, padx=5, pady=5)
list5.bind("<<ComboboxSelected>>", lambda e: update_name_options())

tk.Label(root, text="Packaging:").grid(row=2, column=0, padx=5, pady=5)
list3 = ttk.Combobox(root, state="readonly")
list3.grid(row=2, column=1, padx=5, pady=5)
list3.bind("<<ComboboxSelected>>", lambda e: update_name_options())

#### BOTONES SELECCION DE STACK
tk.Label(root, text="STACK1:").grid(row=2, column=2, padx=5, pady=5)
list6 = ttk.Combobox(root, state="readonly")
list6.grid(row=2, column=3, padx=5, pady=5)

tk.Label(root, text="STACK2:").grid(row=3, column=2, padx=5, pady=5)
list7 = ttk.Combobox(root, state="readonly")
list7.grid(row=3, column=3, padx=5, pady=5)

tk.Label(root, text="STACK3:").grid(row=4, column=2, padx=5, pady=5)
list8 = ttk.Combobox(root, state="readonly")
list8.grid(row=4, column=3, padx=5, pady=5)

tk.Label(root, text="STACK4:").grid(row=5, column=2, padx=5, pady=5)
list9 = ttk.Combobox(root, state="readonly")
list9.grid(row=5, column=3, padx=5, pady=5)

tk.Label(root, text="STACK5:").grid(row=2, column=4, padx=5, pady=5)
list10 = ttk.Combobox(root, state="readonly")
list10.grid(row=2, column=5, padx=5, pady=5)

tk.Label(root, text="STACK6:").grid(row=3, column=4, padx=5, pady=5)
list11 = ttk.Combobox(root, state="readonly")
list11.grid(row=3, column=5, padx=5, pady=5)

tk.Label(root, text="STACK7:").grid(row=4, column=4, padx=5, pady=5)
list12 = ttk.Combobox(root, state="readonly")
list12.grid(row=4, column=5, padx=5, pady=5)

tk.Label(root, text="STACK8:").grid(row=5, column=4, padx=5, pady=5)
list13 = ttk.Combobox(root, state="readonly")
list13.grid(row=5, column=5, padx=5, pady=5)


# Button to take a screenshot
tk.Button(root, text="TOMAR FOTO", command=save_screenshot_and_data).grid(row=6, column=0, columnspan=2, pady=10)

# Status label to display the screenshot filename
status_label = tk.Label(root, text="")
status_label.grid(row=6, column=2, columnspan=2, pady=10)

# Button to take mark End of Stiching
tk.Button(root, text="TOMAR FIN COSIDO", command=save_stiching).grid(row=7, column=0, columnspan=2, pady=10)
# Status label to display End of Stiching
cosido_label = tk.Label(root, text="")
cosido_label.grid(row=7, column=2, columnspan=2, pady=10)

# Start the camera stream in a separate thread
camera_thread = threading.Thread(target=run_camera, daemon=True)
camera_thread.start()

# Initialize the GUI
update_side_and_packing()

# Start the GUI loop
root.mainloop()

# Ensure proper shutdown
capture_flag = False





