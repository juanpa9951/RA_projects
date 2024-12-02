###### DTR USER INPUT- LIVE FEED MEASUREMENT WITH ZOOM AND DIFFERENT COLORS
#### GUI INTERFACE v1 AND MANUAL INPUT AND LOG_DTR


##.....LOAD THE TABLE CALIBRATION DATA.......................................................................................................
import pandas as pd
excel_table_calib = r'C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\feeder_right.xlsx'
Surface_map = pd.read_excel(excel_table_calib, sheet_name='lona_izq', header=0)
tuples_list_real=[]
tuples_list_pixel=[]
for i in range(0,len(Surface_map)):
    x_tup_pixel = Surface_map['X_pixel'][i]
    y_tup_pixel = Surface_map['Y_pixel'][i]
    tuples_list_pixel.append((x_tup_pixel,y_tup_pixel))
    x_tup_real = Surface_map['X_real'][i]
    y_tup_real = Surface_map['Y_real'][i]
    tuples_list_real.append((x_tup_real,y_tup_real))

##........ LOAD DTR DATA...........................................................
excel_DTR = r'C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\DTR.xlsx'
DTR = pd.read_excel(excel_DTR, sheet_name='DTR_sheet', header=0)
stack_list = DTR.iloc[:, 0].tolist()
DTR = DTR.set_index('Name')   #### this is to force the index to be column "Name"

#####...........LOAD PHOTOS PATH.............................
saved_images_path=r"C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\capturas\screenshots"
########.........LOAD DTR REGISTER.............................................................................................................................
dtr_register_path=r"C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\capturas\screenshots"
########.........CAMERA LOG PATH.....................
log_file_path=r"C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\capturas\screenshots\LOG.txt"
log_file_path_DTR=r"C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\capturas\screenshots\LOG_DTR.txt"
def euclidean_distance(point1, point2):
    import math
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)
def find_closest_tupleV2(tuples_list_real,tuples_list_autocad, input_tuple):
    #### same as V1 but only 1 output
    closest_tuple_low = None
    closest_tuple_high = None
    idx_low=None
    idx_high=None

    smallest_distance = float('inf')  # Initialize with a large number
    i=0
    for tup in tuples_list_real:
        distance = euclidean_distance(tup, input_tuple)
        if distance < smallest_distance and tup[0]<input_tuple[0] and tup[1]<input_tuple[1]:
            smallest_distance = distance
            closest_tuple_low = tup
            idx_low=i
        i=i+1

    smallest_distance = float('inf')  # Initialize with a large number
    i=0
    for tup in tuples_list_real:
        distance = euclidean_distance(tup, input_tuple)
        if distance < smallest_distance and tup[0]>input_tuple[0] and tup[1]>input_tuple[1]:
            smallest_distance = distance
            closest_tuple_high = tup
            idx_high=i
        i=i+1

    ### interpolate X value from real to autocad
    target_value=input_tuple[0]
    a_prev = tuples_list_real[idx_low][0]
    a_next = tuples_list_real[idx_high][0]
    b_prev = tuples_list_autocad[idx_low][0]
    b_next = tuples_list_autocad[idx_high][0]
    # Perform linear interpolation
    interpolated_value = b_prev + ((target_value - a_prev) / (a_next - a_prev)) * (b_next - b_prev)
    x_interp=interpolated_value


    ### interpolate Y value from real to autocad
    target_value =input_tuple[1]
    a_prev = tuples_list_real[idx_low][1]
    a_next = tuples_list_real[idx_high][1]
    b_prev = tuples_list_autocad[idx_low][1]
    b_next = tuples_list_autocad[idx_high][1]
    # Perform linear interpolation
    interpolated_value = b_prev + ((target_value - a_prev) / (a_next - a_prev)) * (b_next - b_prev)
    y_interp=interpolated_value

    xy_tup=(x_interp,y_interp)


    return xy_tup,idx_low,idx_high,closest_tuple_low,closest_tuple_high

def distance_real(col1,col2,row1,row2):
    import math
    start_point_pixel = (col1, row1)
    start_point_real, idx_low1, idx_high1, tup_low1, tup_high1 = find_closest_tupleV2(tuples_list_pixel,tuples_list_real,start_point_pixel)
    end_point_pixel = (col2, row2)
    end_point_real, idx_low2, idx_high2, tup_low2, tup_high2 = find_closest_tupleV2(tuples_list_pixel,tuples_list_real,end_point_pixel)
    distance_X=round(abs(end_point_real[0]-start_point_real[0]),2)
    distance_Y=round(abs(end_point_real[1]-start_point_real[1]),2)
    hyp=round(math.sqrt(distance_X**2+distance_Y**2),2)
    return distance_X, distance_Y,hyp

def guardar_registro_dtr(dtr_register_path,stack_name,Width_dtr,Length_dtr,l_real,w_real,L_text,W_text,batch,turno,maquina,quality1,quality2,quality3,quality4,quality5,quality6,item_nr):
    from openpyxl import load_workbook
    from datetime import datetime
    from openpyxl.worksheet.table import Table

    now = datetime.now()
    date=now.date()
    time_now=now.time()

    # Load the workbook and select the worksheet
    file_name_xl='DTR_REGISTER.xlsx'
    file_path = rf"{dtr_register_path}\{file_name_xl}"
    wb = load_workbook(file_path)
    ws = wb.active  # Assuming the table is in the active sheet

    # Find the table by name
    table_name = 'Table1'  # Replace with your table's name
    excel_table = ws.tables[table_name]

    # Get the current table range
    current_range = excel_table.ref  # e.g., "A1:B5"
    start_row = int(current_range.split(':')[1][1:]) + 1  # First empty row below the table
    start_col = current_range.split(':')[0][0]  # First column of the table
    end_col = current_range.split(':')[1][0]  # Last column of the table

    # New data to add
    new_data = [
        [stack_name,Length_dtr,l_real,L_text,Width_dtr,w_real,W_text,quality1,quality2,quality3,quality4,quality5,quality6,batch,turno,maquina,item_nr,time_now,date]
    ]

    # Add new rows to the worksheet
    for i, row_data in enumerate(new_data, start=start_row):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    # Update the table's range to include the new rows
    end_row = start_row + len(new_data) - 1
    excel_table.ref = f"{current_range.split(':')[0]}:{end_col}{end_row}"

    # Save the workbook
    wb.save(file_path)
    print(f"Data added to table '{table_name}' in {file_path}")




import cv2
from datetime import datetime

# Global variables to store the mouse position and clicked points
mouse_x, mouse_y = 0, 0
points = []
lines_drawn = 0
lines = []
line_count = 0
drawing = False
button_pressed = False

# Define button properties
button_text = "Save Screenshot"
button_color = (0, 255, 0)
button_position = (10, 900)
button_size = (150, 30)
button_rect = (button_position[0], button_position[1], button_position[0] + button_size[0], button_position[1] + button_size[1])


def zoom_and_draw_circle(frame, zoom_factor, x, y):
    height, width = frame.shape[:2]

    # Calculate the cropping coordinates
    new_width, new_height = int(width / zoom_factor), int(height / zoom_factor)
    x1 = max(0, x - new_width // 2)
    y1 = max(0, y - new_height // 2)
    x2 = min(width, x + new_width // 2)
    y2 = min(height, y + new_height // 2)

    # Ensure the cropped area does not go out of the frame
    if x2 - x1 < new_width:
        x2 = x1 + new_width
    if y2 - y1 < new_height:
        y2 = y1 + new_height

    # Crop and resize the frame
    zoomed_frame = frame[y1:y2, x1:x2]
    zoomed_frame = cv2.resize(zoomed_frame, (width, height))

    # Draw a small blue circle in the middle of the zoomed frame
    center_x, center_y = width // 2, height // 2
    radius = 7
    color = (255, 0, 0)  # Blue color in BGR
    thickness = -1  # Filled circle
    cv2.circle(zoomed_frame, (center_x, center_y), radius, color, thickness)

    return zoomed_frame


def mouse_callback(event, x, y, flags, param):
    global mouse_x, mouse_y, points, lines_drawn, points, lines, line_count, drawing, button_pressed
    if event == cv2.EVENT_MOUSEMOVE:
        mouse_x, mouse_y = x, y
    if event == cv2.EVENT_LBUTTONDOWN:
        if button_rect[0] <= x <= button_rect[2] and button_rect[1] <= y <= button_rect[3]:
            button_pressed = True
            return

        points.append((x, y))
        drawing = True

        # Once we have two points, draw the line and store the line points
        if len(points) == 2:
            lines.append((points[0], points[1]))
            print(f"Line from {points[0]} to {points[1]}")
            points = []
            line_count += 1
            drawing = False

        # Check if three lines are drawn
        if line_count == 3:
            print("Three lines drawn. Resetting the lines.")
            line_count = 0
            lines.clear()


def main(stack_name,Width_dtr,Length_dtr):
    global mouse_x, mouse_y, points, lines_drawn,button_pressed

    ###### Open a connection to the webcam (0 is the default camera)
    # cap = cv2.VideoCapture("rtsp://LP003:LP003ASM@192.168.2.76:554/stream1")  ### asm001
    # cap = cv2.VideoCapture("rtsp://LP002:LP002ASM@192.168.2.72:554/stream1")  ### asm005
    # cap = cv2.VideoCapture("rtsp://MSM005:LP005ASM@172.16.58.15:554/stream1")  ### ud tapes
    # cap = cv2.VideoCapture("rtsp://LP006:LP006ASM@192.168.2.79:554/stream1")  ### asm003 der
    # cap = cv2.VideoCapture("rtsp://LP005:LP005ASM@192.168.2.75:554/stream1")  ### asm003 izq
    # cap = cv2.VideoCapture("rtsp://LP001:LP001ASM@192.168.2.71:554/stream1")  ### asm002 izq
    # cap = cv2.VideoCapture("rtsp://LP004:LP004ASM@192.168.2.78:554/stream1")  ### asm002 der
    # cap = cv2.VideoCapture("rtsp://LP008:LP008ASM@192.168.2.82:554/stream1") ### asm004 der
    cap = cv2.VideoCapture("rtsp://LP009:LP009ASM@192.168.2.84:554/stream1") ### asm004 izq
    # cap = cv2.VideoCapture("rtsp://RA-camara3:RewAir2023@172.16.58.16:554/stream1")  ## tagging 1
    # cap = cv2.VideoCapture("rtsp://RA-camaras:RewAir2023@172.16.58.142:554/stream1")  ## tagging 1 auxiliar
    # cap = cv2.VideoCapture("rtsp://RA-camara4:RewAir2023@172.16.58.17:554/stream1") ## tagging 2
    # cap = cv2.VideoCapture("rtsp://RA-camara2:RewAir2023@172.16.58.180:554/stream1")  ## tagging 2 auxiliar

    if not cap.isOpened():
        print("Error: Could not open video.")
        return

    cv2.namedWindow('Video Feed')
    cv2.setMouseCallback('Video Feed', mouse_callback)

    zoom_factor = 5.0  #    ### 5.0  Adjust zoom factor as needed
    zoom_size_ratio = 0.3   ### 0.3    # Size of the zoomed window relative to the original frame

    screenshot_count = 0
    #cam_sw=1
    while True:
        # Read a frame from the video feed
        ret, frame = cap.read()

        if not ret:
            print("Error: Could not read frame.")
            break

        height, width = frame.shape[:2]

        # Apply zoom and draw circle to the frame based on the mouse position
        zoomed_frame = zoom_and_draw_circle(frame, zoom_factor, mouse_x, mouse_y)

        # Resize the zoomed frame to fit in the corner
        zoomed_frame_height = int(height * zoom_size_ratio)
        zoomed_frame_width = int(width * zoom_size_ratio)
        zoomed_frame_small = cv2.resize(zoomed_frame, (zoomed_frame_width, zoomed_frame_height))

        # Overlay the zoomed frame onto the original frame
        x_offset, y_offset = 10, 10  # Position of the zoomed frame in the corner
        frame[y_offset:y_offset + zoomed_frame_height, x_offset:x_offset + zoomed_frame_width] = zoomed_frame_small


        # Draw the lines and coordinates on the frame
        color_line1 = (0, 255, 0)
        color_line2 = (255, 150, 0)
        green_color=(0, 255, 0)
        red_color=(0,0,255)

        #### Draw the lines with different colors
        ln=0
        for line in lines:
            if ln == 0:
                color_line = color_line1
            else:
                color_line = color_line2
            cv2.line(frame, line[0], line[1], color_line, 1)
            ln=ln+1

        ###### calculate length and width and comparw with DTR
        for i, line in enumerate(lines):
            if i < 2:  # Only display the coordinates of the first two lines
                col1 = line[0][0]
                col2 = line[1][0]
                row1 = line[0][1]
                row2 = line[1][1]
                xt, yt, hyp = distance_real(col1, col2, row1, row2)
                check_length= (hyp>=Length_dtr-20) and (hyp<=Length_dtr+20)
                check_width= (hyp>=Width_dtr-20) and (hyp<=Width_dtr+20)
                if check_length:
                    L_text="OK"
                else:
                    L_text = "NOT-OK"

                if check_width:
                    W_text="OK"
                else:
                    W_text = "NOT-OK"

                ### assign color to the text
                if i==0:
                    if L_text=="OK":
                      color_text=green_color
                    else:
                      color_text = red_color
                else:
                    if W_text=="OK":
                      color_text=green_color
                    else:
                      color_text = red_color

                ### write the text after drawing the 2 lines already
                if i==0:
                    text1=f"LENGTH = {hyp} {L_text}, dist x= {xt}, dist y= {yt}"
                    cv2.putText(frame,text1, (200, 30 + i * 30), cv2.FONT_HERSHEY_SIMPLEX, 1, color_text, 1, cv2.LINE_AA)
                    l_real=hyp
                else:
                    text2=f"WIDTH = {hyp} {W_text}, dist x= {xt}, dist y= {yt}"
                    cv2.putText(frame,text2, (200, 30 + i * 30), cv2.FONT_HERSHEY_SIMPLEX, 1, color_text, 1, cv2.LINE_AA)
                    w_real=hyp

        # Draw the button
        cv2.rectangle(frame, (button_rect[0], button_rect[1]), (button_rect[2], button_rect[3]), button_color, -1)
        cv2.putText(frame, button_text, (button_rect[0] + 5, button_rect[1] + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5,(0, 0, 0), 1, cv2.LINE_AA)

        # draw name and mesaures from DTR
        cv2.putText(frame, f"STACK {stack_name}, Length_DTR= {Length_dtr}, Width_DTR= {Width_dtr}", (100, 900), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0), 2, cv2.LINE_AA)

        # Display the combined frame
        cv2.imshow('Video Feed', frame)

        # Save screenshot if the button was pressed
        if button_pressed:
            # screenshot_filename = rf"C:\Users\Juan Pablo Lopez\OneDrive - Rewair A S\Documents\Camaras ASM004\capturas\screenshots\{stack_name}_{screenshot_count}.png"
            screenshot_filename = rf"{saved_images_path}\{stack_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            cv2.imwrite(screenshot_filename, frame)
            print(f"\n Foto guardada como  {stack_name}.png")
            button_pressed = False
            break


        #### # Exit loop on 'ESC' key press
        if cv2.waitKey(1) & 0xFF == 27:  # ASCII value for ESC is 27
            break

    # Release the video capture object and close all OpenCV windows
    cap.release()
    cv2.destroyAllWindows()
    return l_real,w_real,L_text,W_text


######  START EXECUTION OF CODE HERE...............................

import tkinter as tk
from tkinter import ttk
from tkinter import Toplevel, messagebox

class AutocompleteCombobox(ttk.Combobox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._full_values = self['values']  # Store the original list of values
        self.bind('<KeyRelease>', self._on_keyrelease)

    def _on_keyrelease(self, event):
        # Get the current input
        typed_text = self.get()

        # Only filter if the length of input is at least 3
        if len(typed_text) < 3:
            self['values'] = self._full_values  # Reset to all options
            return

        # Filter the list based on the typed text
        filtered_values = [
            item for item in self._full_values if typed_text.lower() in item.lower()
        ]

        # Update the combobox values
        self['values'] = filtered_values

        # Automatically show the dropdown if there are filtered values
        if filtered_values:
            self.event_generate('<Down>')

def submit_action(): #### CAMERA FUNCTION
    # Fetch input values
    batch = entry1.get()
    turno = combobox1.get()
    maquina = combobox2.get()
    stack_name = autocomplete_entry.get()
    quality1 = combobox4.get()
    quality2 = combobox5.get()
    quality3 = combobox6.get()
    quality4 = combobox7.get()
    quality5 = combobox8.get()
    quality6 = combobox9.get()
    item_nr  = entry2.get()
    quality_check=0
    if quality1=="Yes" and quality2=="Yes" and quality3=="Yes" and quality4=="Yes" and quality5=="Yes" and quality6=="Yes":
        quality_check=1
    ### Perform some action with the collected data
    stack_check = (DTR.index == stack_name).any()   ### im using stack names as the index of the DTR, here I check if it exists
    if stack_check and quality_check:
        print("\n Activando camara")
        Width_dtr = DTR.at[stack_name, 'Width']
        Length_dtr = DTR.at[stack_name, 'Length']
        ####  execute main code.........................
        l_real,w_real,L_text,W_text=main(stack_name, Width_dtr, Length_dtr)
        print(l_real,w_real,L_text,W_text)
        ######.......... save excel file.................
        guardar_registro_dtr(dtr_register_path, stack_name, Width_dtr, Length_dtr, l_real, w_real, L_text, W_text,batch, turno, maquina,quality1,quality2,quality3,quality4,quality5,quality6,item_nr)
        ######................save log file.......................
        with open(log_file_path, "a") as log_file:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"{timestamp},STACK: {stack_name},BATCH: {batch},MAQUINA: {maquina}, ITEM: {item_nr}\n")
        with open(log_file_path_DTR, "a") as log_file:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"{timestamp},{stack_name},{Length_dtr},{l_real},{L_text},{Width_dtr},{w_real},{W_text},{batch}, {turno}, {maquina},{quality1},{quality2},{quality3},{quality4},{quality5},{quality6},{item_nr}\n")

        result_message="GUARDADO!- Siguiente Stack.."
        print(f"\n", result_message)
        result_label.configure(text=result_message)
        combobox4.set("No")
        combobox5.set("No")
        combobox6.set("No")
        combobox7.set("No")
        combobox8.set("No")
        combobox9.set("No")
    else:
        result_message="Error de Inspecciones"
        print(f"\n", result_message)
        result_label.configure(text=result_message)

def manual_entry(Length_entry, Width_entry,status_label):
    try:
        # Convert entries to floats
        l_real = float(Length_entry.get())
        w_real = float(Width_entry.get())
        batch = entry1.get()
        turno = combobox1.get()
        maquina = combobox2.get()
        stack_name = autocomplete_entry.get()
        quality1 = combobox4.get()
        quality2 = combobox5.get()
        quality3 = combobox6.get()
        quality4 = combobox7.get()
        quality5 = combobox8.get()
        quality6 = combobox9.get()
        item_nr  = entry2.get()
        Width_dtr = DTR.at[stack_name, 'Width']
        Length_dtr = DTR.at[stack_name, 'Length']
        guardar_registro_dtr(dtr_register_path, stack_name, Width_dtr, Length_dtr, l_real, w_real, "", "", batch,
                             turno, maquina, quality1, quality2, quality3, quality4, quality5, quality6, item_nr)
        with open(log_file_path, "a") as log_file:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"{timestamp},STACK: {stack_name},BATCH: {batch},MAQUINA: {maquina}, ITEM: {item_nr}\n")
        with open(log_file_path_DTR, "a") as log_file:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_file.write(f"{timestamp},{stack_name},{Length_dtr},{l_real},{'manual'},{Width_dtr},{w_real},{'manual'},{batch}, {turno}, {maquina},{quality1},{quality2},{quality3},{quality4},{quality5},{quality6},{item_nr}\n")
        status_label.config(text="REGISTRO GUARDADO")
    except:
        status_label.config(text="ERROR NUMERO")

def verificar_manual(stack_name,Length_entry, Width_entry,status_label,status_length,status_width):
    try:
        # Convert entries to floats
        l_real = float(Length_entry.get())
        w_real = float(Width_entry.get())
        Width_dtr = DTR.at[stack_name, 'Width']
        Length_dtr = DTR.at[stack_name, 'Length']
        if l_real <= Length_dtr+20 and l_real >= Length_dtr-20:
            L_text="OK"
        else:
            L_text = "NOT OK"
        if w_real <= Width_dtr+20 and w_real >= Width_dtr-20:
            W_text="OK"
        else:
            W_text = "NOT OK"
        status_length.config(text=L_text)
        status_width.config(text=W_text)
    except:
        status_label.config(text="ERROR NUMERO")

def manual_button():
    stack_name = autocomplete_entry.get()
    Width_dtr = DTR.at[stack_name, 'Width']
    Length_dtr = DTR.at[stack_name, 'Length']

    # Create a new window
    new_window = Toplevel(root)
    new_window.title("Manual Entry")
    new_window.geometry("500x500")  # Width x Height in pixels
    # Add two entry widgets
    tk.Label(new_window, text=f"Stack: {stack_name}").pack(pady=5)
    tk.Label(new_window, text=f"Length DTR: {Length_dtr}").pack(pady=5)
    tk.Label(new_window, text="Length: ").pack(pady=5)
    Length_entry = tk.Entry(new_window)
    Length_entry.pack(pady=5)

    status_length = tk.Label(new_window, text=" ", fg="green")
    status_length.pack(pady=10)


    tk.Label(new_window, text=f"Width DTR: {Width_dtr}").pack(pady=5)
    tk.Label(new_window, text="Width: ").pack(pady=5)
    Width_entry = tk.Entry(new_window)
    Width_entry.pack(pady=5)

    status_width = tk.Label(new_window, text=" ", fg="green")
    status_width.pack(pady=10)

    status_label = tk.Label(new_window, text="Registre datos", fg="green")
    status_label.pack(pady=10)

    # Add a button to execute the function

    tk.Button(
        new_window,
        text="Verificar ",
        command=lambda: verificar_manual(stack_name,Length_entry, Width_entry,status_label,status_length,status_width)
    ).pack(pady=10)

    tk.Button(
        new_window,
        text="Registrar manualmente",
        command=lambda: manual_entry(Length_entry, Width_entry,status_label)
    ).pack(pady=10)




# Create the main window
root = tk.Tk()
root.title("DTR CAMERA")

# Create and place widgets
# Label and Entry for the first input (text input)
label1 = tk.Label(root, text="Batch:")
label1.grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry1 = tk.Entry(root)
entry1.grid(row=0, column=1, padx=10, pady=5)

# Label and Combobox for the second input (list input)
label2 = tk.Label(root, text="Turno:")
label2.grid(row=1, column=0, padx=10, pady=5, sticky="w")
combobox1 = ttk.Combobox(root, values=["Mañana", "Tarde"])
combobox1.grid(row=1, column=1, padx=10, pady=5)
combobox1.set("Mañana")  # Default value

# Label and Combobox for the fourth input (list input)
label3 = tk.Label(root, text="Maquina:")
label3.grid(row=2, column=0, padx=10, pady=5, sticky="w")
combobox2 = ttk.Combobox(root, values=["ASM004", "ASM002", "ASM003"])
combobox2.grid(row=2, column=1, padx=10, pady=5)
combobox2.set("ASM004")  # Default value

# Label and Combobox for the third input (list input)
label4 = tk.Label(root, text="STACK:")
label4.grid(row=4, column=0, padx=10, pady=5, sticky="w")
options = stack_list
autocomplete_entry = AutocompleteCombobox(root, values=options)
autocomplete_entry.grid(row=4, column=1, padx=10, pady=5)


# Additional input
label5 = tk.Label(root, text="Intermediate RM: ")
label5.grid(row=0, column=2, padx=10, pady=5, sticky="w")
combobox4 = ttk.Combobox(root, values=["Yes", "No"])
combobox4.grid(row=0, column=3, padx=10, pady=5)
combobox4.set("No")  # Default value

# Additional input
label6 = tk.Label(root, text="Visual Inspection: ")
label6.grid(row=1, column=2, padx=10, pady=5, sticky="w")
combobox5 = ttk.Combobox(root, values=["Yes", "No"])
combobox5.grid(row=1, column=3, padx=10, pady=5)
combobox5.set("No")  # Default value

# Additional input
label7 = tk.Label(root, text="All marks placed: ")
label7.grid(row=2, column=2, padx=10, pady=5, sticky="w")
combobox6 = ttk.Combobox(root, values=["Yes", "No"])
combobox6.grid(row=2, column=3, padx=10, pady=5)
combobox6.set("No")  # Default value

# Additional input
label8 = tk.Label(root, text="FoD Free: ")
label8.grid(row=0, column=4, padx=10, pady=5, sticky="w")
combobox7 = ttk.Combobox(root, values=["Yes", "No"])
combobox7.grid(row=0, column=5, padx=10, pady=5)
combobox7.set("No")  # Default value

# Additional input
label9 = tk.Label(root, text="Laser Check: ")
label9.grid(row=1, column=4, padx=10, pady=5, sticky="w")
combobox8 = ttk.Combobox(root, values=["Yes", "No"])
combobox8.grid(row=1, column=5, padx=10, pady=5)
combobox8.set("No")  # Default value

# Additional input
label10 = tk.Label(root, text="Overlap Check: ")
label10.grid(row=2, column=4, padx=10, pady=5, sticky="w")
combobox9 = ttk.Combobox(root, values=["Yes", "No"])
combobox9.grid(row=2, column=5, padx=10, pady=5)
combobox9.set("No")  # Default value


# Additional input
label11 = tk.Label(root, text="ITEM:")
label11.grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry2 = tk.Entry(root)
entry2.grid(row=3, column=1, padx=10, pady=5)


# Submit Button
submit_button = tk.Button(root, text="ACTIVAR CAMARA", command=submit_action)
submit_button.grid(row=5, column=5, columnspan=2, pady=10)


# Manual Entry
manual_button = tk.Button(root, text="MANUAL ENTRY", command=manual_button)
manual_button.grid(row=6, column=5, columnspan=2, pady=10)


# Label to display the result message
result_label = tk.Label(root, text=" Camara Lista ", fg="blue", justify="left")
result_label.grid(row=6, column=0, columnspan=2, padx=10, pady=10)


# Run the application
root.mainloop()





